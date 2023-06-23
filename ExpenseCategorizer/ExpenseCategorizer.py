import csv
import os
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

input_filename = input('Enter the name of the input CSV file: ')
output_filename = f"{input_filename} - Categorized.csv"
fieldnames_map = {
    'Transaction Date': ['Posting Date','Transaction Date', 'Date'],
    'Transaction Description': ['Original Description','Description', 'Transaction Description'],
    'Transaction Amount': ['Debit', 'Transaction Amount', 'Amount']
}

categories = defaultdict(list)
with open('Categories.csv') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        categories[row['category']].append(row['keyword'])

def get_field_value(row, field):
    for name in fieldnames_map[field]:
        if name in row:
            return row[name]
    return None

def to_number(value):
    try:
        return float(value)
    except ValueError:
        return value

def categorize(description):
    description = description.lower()
    for category, keywords in categories.items():
        for keyword in keywords:
            if keyword in description:
                return category
    return "Misc"

transactions_by_category = defaultdict(list)
transactions_by_month = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

with open(input_filename) as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        date = datetime.strptime(get_field_value(row, 'Transaction Date'), '%Y-%m-%d')
        month = date.strftime('%Y-%m')
        category = categorize(get_field_value(row, 'Transaction Description'))
        transactions_by_month[month][category][date.day].append(row)

workbook = Workbook()
workbook.remove(workbook.active)
fieldnames = ['Date', 'Desc.'] + list(categories.keys())
totals_row_index = 50

for month, categories in transactions_by_month.items():
    worksheet = workbook.create_sheet(month)
    worksheet.append(fieldnames)
    for day in range(1, 32):
        for category, days in categories.items():
            if day in days:
                for transaction in days[day]:
                    row = [day, get_field_value(transaction, 'Transaction Description')] + [''] * (len(fieldnames) - 2)
                    row[fieldnames.index(category)] = to_number(get_field_value(transaction, 'Transaction Amount'))
                    worksheet.append(row)
    max_length = 0
    for cell in worksheet['B']:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    column_b = worksheet.column_dimensions['B']
    column_b.width = max_length + 2
    totals_row = ['Total', ''] + [''] * (len(fieldnames) - 2)
    for i, category in enumerate(fieldnames[2:]):
        column_letter = get_column_letter(i + 3)
        totals_row[i + 2] = f"=SUM({column_letter}2:{column_letter}{worksheet.max_row})"
    worksheet.insert_rows(totals_row_index)
    worksheet.cell(row=totals_row_index, column=1).value = 'Total'
    for i, value in enumerate(totals_row[2:], start=3):
        worksheet.cell(row=totals_row_index, column=i).value = value

output_filename = f"{input_filename} - Categorized.xlsx"
workbook.save(output_filename)
